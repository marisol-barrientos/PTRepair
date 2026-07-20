# Analysis overlap between requirements presenting violations and those which not (resolution context)
import os
import json
import re
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# ============================================================
# CONFIGURATION
# ============================================================

from pathlib import Path

# Project root
BASE_PATH = Path(__file__).resolve().parents[2]

# Requirements WITH violations
VIOLATION_REQS_DIR = (
    BASE_PATH / "data" / "input" / "compliance_requirements"
)

# Resolution context requirements
RESOLUTION_CONTEXT_DIR = (
    BASE_PATH / "data" / "input" / "resolution_context"
)

# Output directory
OUTPUT_DIR = (
    BASE_PATH / "data" / "output" / "analysis_resolution_context"
)
Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

# ============================================================
# REGEX
# ============================================================

PATTERN_REGEX = re.compile(r'(\w+)\((.*?)\)')


# ============================================================
# HELPERS
# ============================================================

def split_arguments(arg_string):

    args = re.findall(r"'([^']*)'|([^,]+)", arg_string)

    cleaned = []

    for a, b in args:
        val = a if a else b
        cleaned.append(val.strip())

    return cleaned


def extract_elements(requirement_text):

    result = {
        "activities": set(),
        "resources": set(),
        "data": set(),
        "temporal": set(),
    }

    matches = PATTERN_REGEX.findall(requirement_text)

    for pattern_name, arg_string in matches:

        args = split_arguments(arg_string)

        # ----------------------------------------------------
        # ACTIVITY ONLY
        # ----------------------------------------------------
        if pattern_name in {
            "exists",
            "absence",
            "loop",
        }:
            if len(args) >= 1:
                result["activities"].add(args[0])

        # ----------------------------------------------------
        # TEMPORAL
        # ----------------------------------------------------
        elif pattern_name in {
            "directly_follows",
            "leads_to",
            "precedence",
            "leads_to_absence",
            "precedence_absence",
            "parallel",
            "condition_directly_follows",
            "condition_eventually_follows",
        }:
            if len(args) >= 2:
                result["activities"].update([args[0], args[1]])
                result["temporal"].add(pattern_name)

        # ----------------------------------------------------
        # RESOURCE
        # ----------------------------------------------------
        elif pattern_name == "executed_by":

            if len(args) >= 2:
                result["activities"].add(args[0])
                result["resources"].add(args[1])

        # ----------------------------------------------------
        # TIMED
        # ----------------------------------------------------
        elif pattern_name == "timed_alternative":

            if len(args) >= 3:
                result["activities"].update([args[0], args[1]])
                result["temporal"].add("timed_alternative")
                result["temporal"].add(args[2])

        # ----------------------------------------------------
        # DATA
        # ----------------------------------------------------
        elif pattern_name in {
            "activity_sends",
            "activity_receives",
            "data_leads_to_absence",
        }:

            if len(args) >= 2:
                result["activities"].add(args[0])
                result["data"].add(args[1])

    return result


def load_json(filepath):

    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


def get_scenario_name(filename):

    filename = filename.replace(".json", "")
    filename = filename.replace("_req_resolution_context", "")

    return filename


# ============================================================
# LOAD FILES
# ============================================================

violation_files = {
    get_scenario_name(f): os.path.join(VIOLATION_REQS_DIR, f)
    for f in os.listdir(VIOLATION_REQS_DIR)
    if f.endswith(".json")
}

resolution_files = {
    get_scenario_name(f): os.path.join(RESOLUTION_CONTEXT_DIR, f)
    for f in os.listdir(RESOLUTION_CONTEXT_DIR)
    if f.endswith(".json")
}

# ============================================================
# MATCH SCENARIOS
# ============================================================

common_scenarios = sorted(
    set(violation_files.keys())
    & set(resolution_files.keys())
)

# ============================================================
# ANALYSIS
# ============================================================

summary_rows = []

for scenario in common_scenarios:

    print(f"Processing {scenario}")

    violation_path = violation_files[scenario]
    resolution_path = resolution_files[scenario]

    violation_reqs = load_json(violation_path)
    resolution_reqs = load_json(resolution_path)

    violation_elements = defaultdict(set)
    resolution_elements = defaultdict(set)

    # --------------------------------------------------------
    # Extract violation requirements
    # --------------------------------------------------------
    for rid, req_text in violation_reqs.items():

        extracted = extract_elements(req_text)

        for k, v in extracted.items():
            violation_elements[k].update(v)

    # --------------------------------------------------------
    # Extract resolution context requirements
    # --------------------------------------------------------
    for rid, req_text in resolution_reqs.items():

        extracted = extract_elements(req_text)

        for k, v in extracted.items():
            resolution_elements[k].update(v)

    # --------------------------------------------------------
    # Compute overlaps
    # --------------------------------------------------------
    activity_overlap = sorted(
        violation_elements["activities"]
        & resolution_elements["activities"]
    )

    resource_overlap = sorted(
        violation_elements["resources"]
        & resolution_elements["resources"]
    )

    data_overlap = sorted(
        violation_elements["data"]
        & resolution_elements["data"]
    )

    temporal_overlap = sorted(
        violation_elements["temporal"]
        & resolution_elements["temporal"]
    )

    # --------------------------------------------------------
    # Save JSON result
    # --------------------------------------------------------
    per_scenario_result = {
        "activities": activity_overlap,
        "resources": resource_overlap,
        "data": data_overlap,
        "temporal": temporal_overlap,
    }

    json_output = os.path.join(
        OUTPUT_DIR,
        f"{scenario}_overlap_analysis.json"
    )

    with open(json_output, "w", encoding="utf-8") as f:
        json.dump(per_scenario_result, f, indent=4)

    # --------------------------------------------------------
    # Summary row
    # --------------------------------------------------------
    summary_rows.append({
        "scenario": scenario,

        "activity_overlap_count": len(activity_overlap),
        "resource_overlap_count": len(resource_overlap),
        "data_overlap_count": len(data_overlap),
        "temporal_overlap_count": len(temporal_overlap),

        "activity_overlaps": ", ".join(activity_overlap),
        "resource_overlaps": ", ".join(resource_overlap),
        "data_overlaps": ", ".join(data_overlap),
        "temporal_overlaps": ", ".join(temporal_overlap),
    })

# ============================================================
# CREATE DATAFRAME
# ============================================================

df = pd.DataFrame(summary_rows)

# ============================================================
# SAVE EXCEL
# ============================================================

excel_output = os.path.join(
    OUTPUT_DIR,
    "summary_overlap_analysis.xlsx"
)

with pd.ExcelWriter(excel_output, engine="openpyxl") as writer:

    df.to_excel(
        writer,
        sheet_name="Overlap Summary",
        index=False
    )

# ============================================================
# FORMAT EXCEL
# ============================================================

wb = load_workbook(excel_output)
ws = wb["Overlap Summary"]

# Bold headers
for cell in ws[1]:
    cell.font = Font(bold=True)

# Auto column width
for column_cells in ws.columns:

    length = max(
        len(str(cell.value)) if cell.value else 0
        for cell in column_cells
    )

    ws.column_dimensions[
        column_cells[0].column_letter
    ].width = min(length + 5, 60)

wb.save(excel_output)

# ============================================================
# FINISHED
# ============================================================

print("\n================================================")
print("ANALYSIS FINISHED")
print("================================================")

print(f"\nExcel summary saved to:\n{excel_output}")